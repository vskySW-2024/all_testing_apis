from fastapi import FastAPI
from fastapi.responses import StreamingResponse
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from fastapi.middleware.cors import CORSMiddleware
from typing import Dict
import pandas as pd
from openpyxl.utils import get_column_letter

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # Or use specific origins like ["http://localhost:3000"]
    allow_credentials=True,
    allow_methods=["*"],  # Allows all methods (GET, POST, etc.)
    allow_headers=["*"],  # Allows all headers
)

@app.post("/generate-excel/")
async def generate_excel(data: dict):
    # Extract columns and rows from the input JSON
    columns = data.get("columns", [])
    rows = data.get("rows", [])
    sheetname = data.get("sheetname", str)
    filename = data.get("filename", str)

    # Create a DataFrame
    df = pd.DataFrame(rows, columns=columns)

    # Save the DataFrame to an Excel file in memory
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name=sheetname)
        workbook = writer.book
        worksheet = writer.sheets[sheetname]

        # Apply styles to the header row
        header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        header_font = Font(bold=True)

        for cell in worksheet[1]:  # 1st row for the header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = None
    
    buffer.seek(0)

    # Return the Excel file as a streaming response
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}.xlsx"}
    )
@app.post("/generate-mapping-report")
def generate_mapping_report(data: Dict):
    def get_legacy_header(source_column, index=0):
        main_header = []
        current = ""
        for i in source_column:
            header = i[index]
            if not current == header:
                current = header
                main_header.append(header.upper())
            else:
                main_header.append("")
        return main_header
    
    # Create a new Workbook and sheet
    wb = Workbook()
    ws = wb.active
    sheetname = data.get("sheetname", "Sheet1")
    filename = data.get("filename", "report")
    columns = data.get("columns", [])
    rows = data.get("rows", [])
    source_column = [(i["label"].split("-")[0], i["field"].split(".")[0], i["field"].split(".")[1]) for i in columns]

    # Rename the active sheet
    ws.title = sheetname

    # Write the headers and rows
    ws.append(get_legacy_header(source_column, 0))
    fill = PatternFill(start_color="3a6c97", end_color="3a6c97", fill_type="solid")
    font = Font(bold=True, color="ffffff")
    for cell in ws[1]:
            cell.fill = fill
            cell.font = font

    ws.append(get_legacy_header(source_column, 1))
    for cell in ws[2]:
            cell.fill = fill
            cell.font = font

    ws.append([i[2].upper() for i in source_column])
    for cell in ws[3]:
            cell.fill = fill
            cell.font = font

    blue_fill = PatternFill(start_color="1b75ab", end_color="1b75ab", fill_type="solid")  # Blue background
    source_num=get_column_letter(len([x for x in source_column if x[0].lower() == "source"]))
    for row in ws[f"A1:{source_num}3"]:  # Specify the range
        for cell in row:
            cell.fill = blue_fill

    for row in rows:
        i_row = []
        for column in source_column:
            j_column = f"{column[1]}.{column[2]}"
            i_row.append(row.get(column[1], {}).get(column[2], ""))
        ws.append(i_row)
    
    # Save the workbook to an in-memory buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Return the Excel file as a StreamingResponse
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}.xlsx"}
    )


# Run the server with: uvicorn filename:app --reload
